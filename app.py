import os
import sqlite3
import logging
import random
import string
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, jsonify, send_file
)
from flask_mail import Mail, Message
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from flask_caching import Cache
import bcrypt
from dotenv import load_dotenv

# =========================
# .env Yükle
# =========================
load_dotenv()

# =========================
# Uygulama
# =========================
app = Flask(__name__)

# Production statik servis (opsiyonel)
if os.environ.get('FLASK_ENV') == 'production':
    try:
        from whitenoise import WhiteNoise
        app.wsgi_app = WhiteNoise(app.wsgi_app, root='static/')
        app.wsgi_app.add_files('static/', prefix='static/')
    except ImportError:
        pass

# Secret key
app.secret_key = os.environ.get('SECRET_KEY', 'dev_secret')

# Debug
app.config['DEBUG'] = os.environ.get('FLASK_ENV') != 'production'

# Rate limit
limiter = Limiter(get_remote_address, app=app, default_limits=["200 per day", "50 per hour"], storage_uri="memory://")

# Cache
cache = Cache(app, config={'CACHE_TYPE': 'simple', 'CACHE_DEFAULT_TIMEOUT': 60})

# Logger
app.logger.setLevel(logging.INFO)

# Mail
app.config['MAIL_SERVER'] = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.environ.get('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = os.environ.get('MAIL_USE_TLS', 'true').lower() == 'true'
app.config['MAIL_USERNAME'] = os.environ.get('MAIL_USERNAME', '')
app.config['MAIL_PASSWORD'] = os.environ.get('MAIL_PASSWORD', '')
app.config['MAIL_DEFAULT_SENDER'] = os.environ.get('MAIL_DEFAULT_SENDER', '')
EMAIL_ENABLED = os.environ.get('EMAIL_ENABLED', 'false').lower() == 'true'
mail = Mail(app)

# DB
DATABASE = os.environ.get('DATABASE_PATH', 'matbaa_takip.db')

# Session süresi
app.permanent_session_lifetime = timedelta(hours=2)

# Security headers (prod)
if os.environ.get('FLASK_ENV') == 'production':
    @app.after_request
    def add_security_headers(response):
        response.headers['X-Content-Type-Options'] = 'nosniff'
        response.headers['X-Frame-Options'] = 'DENY'
        response.headers['X-XSS-Protection'] = '1; mode=block'
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
        return response

# =========================
# DB Yardımcıları
# =========================
def get_db_connection():
    try:
        conn = sqlite3.connect(DATABASE, timeout=10.0)
        conn.row_factory = sqlite3.Row
        conn.execute('PRAGMA foreign_keys = ON')
        return conn
    except Exception as e:
        app.logger.error(f"DB bağlantı hatası: {e}")
        return None

def hash_password(password: str) -> str:
    salt = bcrypt.gensalt()
    return bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def init_db():
    """Tabloları oluştur ve admin'i hazırla"""
    conn = get_db_connection()
    if not conn:
        return
    cur = conn.cursor()

    # books
    cur.execute("""
        CREATE TABLE IF NOT EXISTS books (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            author_name TEXT NOT NULL,
            order_quantity INTEGER NOT NULL,
            size TEXT NOT NULL,
            status TEXT DEFAULT 'Hazırlanıyor',
            track_code TEXT UNIQUE NOT NULL,
            customer_email TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute('CREATE INDEX IF NOT EXISTS idx_track_code ON books(track_code)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_status ON books(status)')
    cur.execute('CREATE INDEX IF NOT EXISTS idx_created_at ON books(created_at)')

    # users
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT DEFAULT 'admin',
            is_active BOOLEAN DEFAULT 1,
            last_login TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # contact messages
    cur.execute("""
        CREATE TABLE IF NOT EXISTS contact_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            email TEXT NOT NULL,
            message TEXT NOT NULL,
            is_read BOOLEAN DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # Admin .env'den
    admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
    admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')

    cur.execute("DELETE FROM users WHERE username = ?", (admin_username,))
    cur.execute("INSERT INTO users (username, password) VALUES (?, ?)",
                (admin_username, hash_password(admin_password)))

    conn.commit()
    conn.close()
    app.logger.info(f"✅ Admin oluşturuldu: {admin_username}")

init_db()  # app start’ta bir kez

# =========================
# Util
# =========================
def sanitize_input(s: str) -> str:
    return (s or '').strip()

def generate_track_code() -> str:
    # 3 harf + 6 rakam
    return ''.join(random.choices(string.ascii_uppercase, k=3)) + ''.join(random.choices(string.digits, k=6))

def get_unique_track_code() -> str:
    """Veritabanında olmayan benzersiz takip kodu üret"""
    conn = get_db_connection()
    if not conn:
        return generate_track_code()

    cur = conn.cursor()
    for _ in range(20):
        code = generate_track_code()
        cur.execute("SELECT 1 FROM books WHERE track_code = ?", (code,))
        if not cur.fetchone():
            conn.close()
            return code
    conn.close()
    return generate_track_code()

def send_email_notification(to_email: str, subject: str, body: str):
    if not EMAIL_ENABLED:
        return False
    try:
        msg = Message(subject=subject, recipients=[to_email], body=body)
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Email gönderme hatası: {e}")
        return False

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            flash('Lütfen giriş yapın', 'error')
            return redirect(url_for('login'))
        # timeout
        last = session.get('last_activity')
        now = datetime.now()
        if last:
            try:
                if now - datetime.fromisoformat(last) > app.permanent_session_lifetime:
                    session.clear()
                    flash('Oturum zaman aşımına uğradı', 'warning')
                    return redirect(url_for('login'))
            except Exception:
                pass
        session['last_activity'] = now.isoformat()
        return f(*args, **kwargs)
    return decorated

# =========================
# Routes — Public
# =========================
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/health')
def health():
    ok = True
    try:
        conn = get_db_connection()
        if conn:
            conn.execute("SELECT 1")
            conn.close()
    except Exception:
        ok = False
    return jsonify({"status": "ok" if ok else "fail", "ts": datetime.now().isoformat()})

@app.route('/track', methods=['GET', 'POST'])
def track():
    if request.method == 'POST':
        track_code = sanitize_input(request.form.get('track_code'))
        if not track_code:
            flash('Lütfen takip kodu girin', 'error')
            return render_template('track.html')

        conn = get_db_connection()
        if conn:
            cur = conn.cursor()
            cur.execute("SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at FROM books WHERE track_code = ?", (track_code,))
            row = cur.fetchone()
            conn.close()
            if row:
                book = dict(row)
                return render_template('track.html', book=book, found=True)
            else:
                flash('Takip kodu bulunamadı', 'error')
    return render_template('track.html')

@app.route('/contact', methods=['POST'])
@limiter.limit("3 per minute")
def contact():
    name = sanitize_input(request.form.get('name'))
    email = sanitize_input(request.form.get('email'))
    message = sanitize_input(request.form.get('message'))
    if not name or not email or not message:
        return jsonify({'success': False, 'message': 'Tüm alanları doldurun'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'message': 'Veritabanı bağlantı hatası'}), 500
    cur = conn.cursor()
    cur.execute("INSERT INTO contact_messages (name, email, message) VALUES (?, ?, ?)", (name, email, message))
    conn.commit()
    conn.close()

    # İsteğe bağlı e-posta
    if EMAIL_ENABLED:
        try:
            send_email_notification(
                to_email=os.environ.get('MAIL_USERNAME', ''),
                subject=f"Yeni İletişim Mesajı - {name}",
                body=f"Ad: {name}\nE-posta: {email}\nMesaj: {message}\nTarih: {datetime.now():%d.%m.%Y %H:%M}"
            )
        except Exception:
            pass

    return jsonify({'success': True, 'message': 'Mesajınız alındı'})

# =========================
# Auth
# =========================
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def login():
    if request.method == 'POST':
        username = sanitize_input(request.form.get('username'))
        password = request.form.get('password') or ''
        if not username or not password:
            flash('Kullanıcı adı ve şifre zorunlu', 'error')
            return render_template('login.html')

        conn = get_db_connection()
        if not conn:
            flash('Veritabanı hatası', 'error')
            return render_template('login.html')
        cur = conn.cursor()
        cur.execute("SELECT id, username, password FROM users WHERE username = ?", (username,))
        user = cur.fetchone()
        conn.close()

        if user and verify_password(password, user['password']):
            session.permanent = True
            session['admin_logged_in'] = True
            session['admin_username'] = user['username']
            session['admin_user_id'] = user['id']
            session['last_activity'] = datetime.now().isoformat()
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Kullanıcı adı veya şifre hatalı', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Çıkış yapıldı', 'success')
    return redirect(url_for('login'))

# =========================
# Admin — Dashboard
# =========================
@app.route('/admin/dashboard')
@login_required
@cache.cached(timeout=60)
def admin_dashboard():
    conn = get_db_connection()
    total_books = active_orders = completed_orders = 0
    recent = []
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM books")
        total_books = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM books WHERE status != 'Hazır'")
        active_orders = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM books WHERE status = 'Hazır'")
        completed_orders = cur.fetchone()[0]
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email
                       FROM books ORDER BY created_at DESC LIMIT 5""")
        recent = [dict(row) for row in cur.fetchall()]
        conn.close()
    return render_template('admin_dashboard.html',
                           total_books=total_books,
                           active_orders=active_orders,
                           completed_orders=completed_orders,
                           recent_books=recent)

# =========================
# Admin — Kitaplar
# =========================
@app.route('/admin/books')
@login_required
def admin_books_all():
    conn = get_db_connection()
    books = []
    if conn:
        cur = conn.cursor()
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                       FROM books ORDER BY created_at DESC""")
        books = [dict(row) for row in cur.fetchall()]
        conn.close()
    return render_template('books_list.html', books=books, title="Tüm Siparişler", type="all")

@app.route('/admin/books/active')
@login_required
def admin_books_active():
    conn = get_db_connection()
    books = []
    if conn:
        cur = conn.cursor()
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                       FROM books WHERE status != 'Hazır' ORDER BY created_at DESC""")
        books = [dict(row) for row in cur.fetchall()]
        conn.close()
    return render_template('books_list.html', books=books, title="Aktif Siparişler", type="active")

@app.route('/admin/books/completed')
@login_required
def admin_books_completed():
    conn = get_db_connection()
    books = []
    if conn:
        cur = conn.cursor()
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                       FROM books WHERE status = 'Hazır' ORDER BY created_at DESC""")
        books = [dict(row) for row in cur.fetchall()]
        conn.close()
    return render_template('books_list.html', books=books, title="Tamamlanan Siparişler", type="completed")

@app.route('/admin/add', methods=['GET', 'POST'])
@login_required
def add_book():
    if request.method == 'POST':
        title = sanitize_input(request.form.get('title'))
        author_name = sanitize_input(request.form.get('author_name'))
        order_quantity = (request.form.get('order_quantity') or '').strip()
        size = sanitize_input(request.form.get('size'))
        status = sanitize_input(request.form.get('status')) or 'Sipariş Alındı'
        track_code = sanitize_input(request.form.get('track_code')) or get_unique_track_code()
        customer_email = sanitize_input(request.form.get('customer_email'))
        send_email = request.form.get('send_email') == 'on'

        # basit validasyon
        errors = []
        if not title: errors.append('Başlık zorunlu')
        if not author_name: errors.append('Yazar zorunlu')
        if not order_quantity.isdigit(): errors.append('Adet sayısal olmalı')
        if not size: errors.append('Ebat zorunlu')
        if errors:
            for e in errors: flash(e, 'error')
            return render_template('add_book.html', preset={'title': title, 'author_name': author_name,
                                                            'order_quantity': order_quantity, 'size': size,
                                                            'status': status, 'track_code': track_code,
                                                            'customer_email': customer_email})
        conn = get_db_connection()
        if not conn:
            flash('DB bağlantı hatası', 'error')
            return render_template('add_book.html')

        cur = conn.cursor()
        try:
            cur.execute("""INSERT INTO books (title, author_name, order_quantity, size, status, track_code, customer_email)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (title, author_name, int(order_quantity), size, status, track_code, customer_email or None))
            conn.commit()
            flash('Kayıt eklendi', 'success')

            if send_email and customer_email:
                try:
                    send_email_notification(customer_email,
                                            subject='Siparişiniz oluşturuldu',
                                            body=f"{title} ({track_code}) siparişiniz alınmıştır.")
                except Exception:
                    pass

            return redirect(url_for('admin_books_all'))
        except sqlite3.IntegrityError:
            flash('Takip kodu zaten mevcut, yenisini deneyin', 'error')
        except Exception as e:
            app.logger.error(f'Insert hatası: {e}')
            flash('Kayıt eklenirken hata', 'error')
        finally:
            conn.close()

    # GET
    return render_template('add_book.html', suggested_code=get_unique_track_code())

# =========================
# Admin — Excel Rapor
# =========================
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_excel_report(books, title="Matbaa Raporu"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapor"

    # Başlık
    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Head
    headers = ['Başlık', 'Yazar', 'Adet', 'Ebat', 'Durum', 'Takip Kodu', 'E-posta']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        c.alignment = Alignment(horizontal='center')

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    if books:
        for r, b in enumerate(books, 4):
            ws.cell(r, 1, b.get('title', '')).border = thin
            ws.cell(r, 2, b.get('author_name', '')).border = thin
            ws.cell(r, 3, b.get('order_quantity', '')).border = thin
            ws.cell(r, 4, b.get('size', '')).border = thin
            ws.cell(r, 5, b.get('status', '')).border = thin
            ws.cell(r, 6, b.get('track_code', '')).border = thin
            ws.cell(r, 7, b.get('customer_email', '')).border = thin
    else:
        ws.merge_cells('A4:G4')
        ws['A4'] = "Veri bulunamadı"
        ws['A4'].alignment = Alignment(horizontal='center')

    widths = [30, 20, 10, 15, 15, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@app.route('/admin/update/<int:book_id>', methods=['GET', 'POST'])
@login_required
def update_book(book_id):
    if request.method == 'POST':
        title = sanitize_input(request.form.get('title', ''))
        author_name = sanitize_input(request.form.get('author_name', ''))
        order_quantity = request.form.get('order_quantity')
        size = sanitize_input(request.form.get('size', ''))
        status = request.form.get('status', 'Hazırlanıyor')
        customer_email = sanitize_input(request.form.get('customer_email', ''))
        
        conn = get_db_connection()
        if conn:
            cur = conn.cursor()
            cur.execute("""
                UPDATE books SET title=?, author_name=?, order_quantity=?, size=?, status=?, customer_email=?, updated_at=CURRENT_TIMESTAMP
                WHERE id=?
            """, (title, author_name, order_quantity, size, status, customer_email, book_id))
            conn.commit()
            conn.close()
            flash('Kitap başarıyla güncellendi.', 'success')
            return redirect(url_for('admin_dashboard'))
    
    # GET request - kitap bilgilerini getir
    conn = get_db_connection()
    if conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM books WHERE id = ?", (book_id,))
        book = cur.fetchone()
        conn.close()
        if book:
            return render_template('update_book.html', book=dict(book))
    flash('Kitap bulunamadı.', 'error')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/delete/<int:book_id>', methods=['DELETE'])
@login_required
def delete_book(book_id):
    conn = get_db_connection()
    if conn:
        cur = conn.cursor()
        cur.execute("DELETE FROM books WHERE id = ?", (book_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': 'Kitap başarıyla silindi'})
    return jsonify({'success': False, 'error': 'Veritabanı bağlantı hatası'})

@app.route('/admin/export-excel/<report_type>')
@login_required
def export_excel_report(report_type):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'DB bağlantı hatası'}), 500
    cur = conn.cursor()

    if report_type == 'all':
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                       FROM books ORDER BY created_at DESC""")
    elif report_type == 'active':
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                       FROM books WHERE status != 'Hazır' ORDER BY created_at DESC""")
    elif report_type == 'completed':
        cur.execute("""SELECT id, title, author_name, order_quantity, size, status, track_code, customer_email, created_at
                       FROM books WHERE status = 'Hazır' ORDER BY created_at DESC""")
    else:
        conn.close()
        return jsonify({'error': 'Geçersiz rapor türü'}), 400

    books = [dict(row) for row in cur.fetchall()]
    conn.close()

    buf = generate_excel_report(books, f"Matbaa Raporu - {report_type.title()}")
    filename = f"matbaa_rapor_{report_type}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return send_file(buf, as_attachment=True,
                     download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# =========================
# Admin — İletişim Mesajları
# =========================
@app.route('/admin/contact-messages')
@login_required
def contact_messages():
    conn = get_db_connection()
    msgs = []
    if conn:
        cur = conn.cursor()
        cur.execute("""SELECT id, name, email, message, is_read, created_at
                       FROM contact_messages ORDER BY created_at DESC""")
        msgs = [dict(row) for row in cur.fetchall()]
        conn.close()
    return render_template('contact_messages.html', contact_messages=msgs)

@app.route('/admin/contact-messages/<int:message_id>')
@login_required
def get_contact_message(message_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB hatası'}), 500
    cur = conn.cursor()
    cur.execute("""SELECT id, name, email, message, is_read, created_at
                   FROM contact_messages WHERE id = ?""", (message_id,))
    row = cur.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'error': 'Mesaj bulunamadı'}), 404
    data = dict(row)
    # tarih formatla
    try:
        if isinstance(data['created_at'], str):
            dt = datetime.strptime(data['created_at'], '%Y-%m-%d %H:%M:%S')
            data['created_at'] = dt.strftime('%d.%m.%Y %H:%M')
    except Exception:
        pass
    return jsonify({'success': True, 'message': data})

@app.route('/admin/contact-messages/<int:message_id>/read', methods=['POST'])
@login_required
def mark_message_read(message_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'DB hatası'}), 500
    cur = conn.cursor()
    cur.execute("UPDATE contact_messages SET is_read = 1 WHERE id = ?", (message_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/admin/contact-messages/<int:message_id>', methods=['DELETE'])
@login_required
def delete_contact_message(message_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'DB hatası'}), 500
    cur = conn.cursor()
    cur.execute("DELETE FROM contact_messages WHERE id = ?", (message_id,))
    conn.commit()
    ok = cur.rowcount > 0
    conn.close()
    return jsonify({'success': ok}) if ok else (jsonify({'error': 'Bulunamadı'}), 404)

@app.route('/admin/contact-messages/mark-all-read', methods=['POST'])
@login_required
def mark_all_messages_read():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'DB hatası'}), 500
    cur = conn.cursor()
    cur.execute("UPDATE contact_messages SET is_read = 1 WHERE is_read = 0")
    conn.commit()
    count = cur.rowcount
    conn.close()
    return jsonify({'success': True, 'updated_count': count})

@app.route('/admin/contact-messages/delete-all', methods=['DELETE'])
@login_required
def delete_all_messages():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'DB hatası'}), 500
    cur = conn.cursor()
    cur.execute("DELETE FROM contact_messages")
    conn.commit()
    count = cur.rowcount
    conn.close()
    return jsonify({'success': True, 'deleted_count': count})

# =========================
# Hata Sayfaları
# =========================
@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    if os.environ.get('FLASK_ENV') == 'production':
        return render_template('errors/500.html'), 500
    return f"<h1>Internal Server Error</h1><pre>{error}</pre>", 500

@app.errorhandler(429)
def rate_limit_error(e):
    return jsonify({'error': 'Çok fazla istek. Biraz bekleyin.'}), 429

# =========================
# Çalıştırma
# =========================
def create_app():
    return app

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=app.config['DEBUG'])
